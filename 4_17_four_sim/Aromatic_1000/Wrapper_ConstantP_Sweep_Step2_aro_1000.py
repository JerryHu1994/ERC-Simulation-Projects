"""This is the main wrapper function.  The two from imports refer the function"""

from Wiebe_two_zone_crev_heatxfer_reformate_LLNL_heatrel_HTCONDOR import Engine


import sys
import numpy as np
import cantera as ct
#import matplotlib.pyplot as plt
import csv
import datetime
import time
import h5py
# from scipy.interpolate import interp1d 
import math
from openpyxl import load_workbook

"""Evaluate the temperature range"""        
         
if (len(sys.argv) != 2):
    print("Usage: python Wrapper_constantmass_Sweep_CONDOR.py jobid")

job_id = (int)(sys.argv[1])

ivc_range=np.linspace(310,440,27)        
         
n = ivc_range[job_id]
print("The temperature is {}".format(n))

#for n in np.nditer(ivc_range):
tic=time.time()
fname='Aromatic_1000_121C_CONDOR_step2.xlsx'

wb=load_workbook(fname)
name=datetime.datetime.now().strftime("20%y%m%d_%H-%M.%S_Aro_1000rpm_Sweep_Step2")
ws=wb.active
ws['A16']=int(n)
ws['A29']=1 #turn end gas chemistry on/off
    #ws['A27']=0 #Nu Constant
ws['A21']=1 #turn main combustion on/off
#    P_ivc=((0.000939377)*(275.772)*(int(n)))/(0.00052138)/1e5 #Recalculate Pressure to maintain Mass
#    ws['A17']=P_ivc
wb.save(name+'_'+str(n)+'.xlsx')     #---Make sure to turn this on for save
[TEST2, SPEC2]=Engine(name+'_'+str(n)+'.xlsx')
np.savez(name+'_'+str(n),outdat=TEST2, species=SPEC2)        #------Make sure to turn this on for save

"""Saving as a zip file to get both arrays"""
