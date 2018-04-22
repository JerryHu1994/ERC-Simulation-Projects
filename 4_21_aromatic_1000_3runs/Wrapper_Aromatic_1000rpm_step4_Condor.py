"""This is the main wrapper function.  The two from imports refer the function"""

from Wiebe_two_zone_crev_heatxfer_reformate_LLNL_heatrel_HTCONDOR import Engine
#from Wiebe_two_zone_crev_heatxfer_reformate_ERCPRF_heatrel_egfix_CONDOR import Engine
#from Wiebe_two_zone_crev_heatxfer_reformate_ERCPRF_heatrel_egfix_species import Engine

import sys
import numpy as np
import cantera as ct
#import matplotlib.pyplot as plt
import csv
import datetime
import time
import h5py
# from scipy.interpolate import interp1d 
import math
from openpyxl import load_workbook

"""Evaluate the temperature range"""        
         
ivc_range_arr = np.array([375, 375, 375])
Wiebe_theta_0_arr = np.array([13, 13, 13])
Wiebe_delta_theta_arr = np.array([37, 40, 43]) 
BFlag_arr = np.array([0.95, 0,95, 0.95])


if (len(sys.argv) != 2):
    print("Usage: python Wrapper_constantmass_Sweep_CONDOR.py jobid")

job_id = (int)(sys.argv[1])
print("The job number is {}".format(job_id))

n = ivc_range_arr[job_id]
Wiebe_theta_0=Wiebe_theta_0_arr[job_id]
Wiebe_delta_theta=Wiebe_delta_theta_arr[job_id]
BFlag=BFlag_arr[job_id]

#for n in np.nditer(ivc_range):
tic=time.time()
fname='Aro_1000_step4_CONDOR.xlsx'

wb=load_workbook(fname)
name=datetime.datetime.now().strftime("20%y%m%d_%H-%M.%S_Aro_1000rpm_step4_Wiebe")
ws=wb.active
ws['A13']=Wiebe_theta_0
ws['A14']=Wiebe_delta_theta
ws['A16']=int(n)
ws['A22']=BFlag
ws['A29']=1 #turn end gas chemistry on/off
    #ws['A27']=0 #Nu Constant
ws['A21']=1 #turn main combustion on/off
P_ivc=((9.145973697404853e-04)*(275.762)*(int(n)))/(0.00052138)/1e5 #Recalculate Pressure to maintain Mass
ws['A17']=P_ivc
wb.save(name+'_'+str(n)+'.xlsx')     #---Make sure to turn this on for save
[TEST2, SPEC2]=Engine(name+'_'+str(n)+'.xlsx')
np.savez(name+'_'+str(n),outdat=TEST2, species=SPEC2)        #------Make sure to turn this on for save

"""Saving as a zip file to get both arrays"""
