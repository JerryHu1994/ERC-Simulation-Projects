def CV_IgDelay_Thesis(x_initial,Temp,Press,PureFlag,fname,RK_Flag):
    import sys
    import time as ttime
    import numpy as np
    import cantera as ct
    #import matplotlib.pyplot as plt
    import csv
    #from scipy.interpolate import interp1d 
    import math
    from openpyxl import load_workbook
    ct.suppress_thermo_warnings()
    
    if RK_Flag==1:
        gas=ct.Solution('renKokjohn.cti')
    else:    
        gas = ct.Solution('LLNL_gasoline_20170621_nox_galway.cti')
    
    if PureFlag==1 and RK_Flag!=1:
        """=========Determine Species if using Pure Fuel=============="""
        #gas1 = ct.Solution('ic8_ver3_mech.cti')
        #gas = ct.Solution('RenKokjohn.cti')
        Phi=1
        x_init=np.zeros(gas.X.shape)
        wb=load_workbook(fname)
        ws=wb.active
        c4=ws['C41'].value #0.01      #n-butane
        c5=ws['C42'].value #0.0      #iso-pentane
        c5_2=ws['C43'].value #0.04   #n-pentane
        ic8=ws['C44'].value #0.93    #Iso-octane
        c6=ws['C45'].value #0.0      #added for 1-hexene
        nc7=ws['C46'].value #0.0    #n-heptane
        c7_2=ws['C47'].value #0.0    #tolune
        tmb=ws['C48'].value #0.02     #1,2,4 Trimethyl benzene 
        eth=ws['C49'].value #0       #Ethanol      
        
        
        
        x_init[gas.species_index('C4H10')]=c4
        x_init[gas.species_index('IC5H12')]=c5
        x_init[gas.species_index('NC5H12')]=c5_2
        x_init[gas.species_index('C6H12-1')]=c6
        x_init[gas.species_index('NC7H16')]=nc7
        x_init[gas.species_index('C6H5CH3')]=c7_2
        x_init[gas.species_index('IC8')]=ic8
        x_init[gas.species_index('T124MBZ')]=tmb
        x_init[gas.species_index('C2H5OH')]=eth
        #    x_init[gas.species_index('NO')]=150e-6  #150 ppm NO
               
        ## Determine global composition
        carbon=np.zeros(gas.X.size); 
        hydrogen=np.zeros(gas.X.size);
        oxygen=np.zeros(gas.X.size)
        for i in range(gas.n_species):
            carbon[i]=gas.n_atoms(i,'C')
            hydrogen[i]=gas.n_atoms(i,'H')
            oxygen[i]=gas.n_atoms(i,'O')
            nC=np.dot(x_init,carbon)
            nH=np.dot(x_init,hydrogen)
            nO=np.dot(x_init,oxygen)
        
        print(nC)
        print(nH)
        print(nO)
        
        x_init[gas.species_index('O2')]=(nC+nH/4-nO/2)/Phi
        x_init[gas.species_index('N2')]=3.76*(nC+nH/4-nO/2)/Phi
    elif PureFlag==1 and RK_Flag==1:
        """=========Determine Species if using Pure Fuel=============="""
        #gas1 = ct.Solution('ic8_ver3_mech.cti')
        #gas = ct.Solution('RenKokjohn.cti')
        Phi=1
        x_init=np.zeros(gas.X.shape)
        wb=load_workbook(fname)
        ws=wb.active
        c4=ws['C50'].value #0.01      #n-butane
        c5=ws['C51'].value #0.0      #iso-pentane
        c5_2=ws['C52'].value #0.04   #n-pentane
        ic8=ws['C53'].value #0.93    #Iso-octane
        c6=ws['C54'].value #0.0      #added for 1-hexene
        nc7=ws['C55'].value #0.0    #n-heptane
        c7_2=ws['C56'].value #0.0    #tolune
        tmb=ws['C57'].value #0.02     #1,2,4 Trimethyl benzene 
        eth=ws['C58'].value #0       #Ethanol
        
        x_init[gas.species_index('nC7h16')]=nc7
        x_init[gas.species_index('jc8h16')]=c6 #added for di-isobutylene
        x_init[gas.species_index('ic8h18')]=ic8
        x_init[gas.species_index('C2H5OH')]=eth
        x_init[gas.species_index('c7h8')]=c7_2
        
        carbon=np.zeros(gas.X.size); 
        hydrogen=np.zeros(gas.X.size);
        oxygen=np.zeros(gas.X.size)
        for i in range(gas.n_species):
            carbon[i]=gas.n_atoms(i,'C')
            hydrogen[i]=gas.n_atoms(i,'H')
            oxygen[i]=gas.n_atoms(i,'O')
            nC=np.dot(x_init,carbon)
            nH=np.dot(x_init,hydrogen)
            nO=np.dot(x_init,oxygen)
            
        x_init[gas.species_index('O2')]=(nC+nH/4-nO/2)/Phi
        x_init[gas.species_index('N2')]=3.76*(nC+nH/4-nO/2)/Phi
        
#        print(nC)
#        print(nH)
#        print(nO)
    else:
         x_init=x_initial
    
#    print('Starting Reactor')
#    print('whatever')        
    gas.TPX=Temp,Press,x_init
#    print(gas.report())     
    r1=ct.IdealGasReactor(gas)     #Ignition Delay reactor    
    sim2= ct.ReactorNet([r1])    
    time=0      #Initialize Time
    Temp_cur=0#ig_temp_cont[n]
    Press_cv=[]
    Temp_cv=[]
    timeapp=[]
#    print('Starting Calculation')
#    ttime.sleep(5)
    
    while (Temp_cur < Temp+50 and time < 20e-3):
        time += 1.e-6
        sim2.advance(time)
        time_cur=time #milliseconds
        Reac_Press=r1.thermo.P
        Reac_Temp=r1.T
        Temp_cur=Reac_Temp
#            times.append(time_cur)
        Press_cv.append(Reac_Press)
        Temp_cv.append(Reac_Temp)
        timeapp.append(time)
#        if np.remainder(time,1.e-3)<1e-8:
#            print(time)
#            print('Running')
    Final_temp=Reac_Temp     
    tau=(time_cur)
            
    return tau, Final_temp       
    