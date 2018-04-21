def Engine(fname):
    import sys
    import numpy as np
    import cantera as ct
    import matplotlib.pyplot as plt
    import csv
    import time
    # from scipy.interpolate import interp1d 
    import math
    from openpyxl import load_workbook

    def volume(t):
        theta=t*omega+IVC*np.pi/180 #theta in radian; TDC corresponds to theta=0
        v=V_min*(1+((r_c-1)/2)*(L/a + 1 - np.cos(theta) - np.sqrt( pow(L/a,2)-pow(np.sin(theta),2) ) ) )
        return v
     
    def surfA(t):
        area=np.pi*math.pow(D,2)/4
        SA=2*area+np.pi*D*volume(t)/area
        return SA
    
    def vel(t): #Note - the only input to the passed through function is time
        theta=t*omega+IVC*np.pi/180 #TDC corresponds to theta=0
        z=np.sqrt(pow(L/a,2)-pow(np.sin(theta),2))
        v=omega*(V_min/A_p)*((r_c-1)/2)*np.sin(theta)*(1+np.cos(theta)/z)
        return v
    
    def qfluxB(t): #Note - the only input to the passed through function is time
        rho=m_t/volume(t)
        T=(unb.mass*unb.thermo.T+bur.mass*bur.thermo.T)/(unb.mass+bur.mass)
        k=k_0*pow(T,n_k)        #W/m-K, for air
        mu=mu_0*pow(T,n_mu)      #kg/m-s, for air
        Re=rho*MPS*D/mu
        Nu=Nu_0*pow(Re,n_Nu)
        h=Nu*k/D
        area=surfA(t)*bur.volume/volume(t)
        q=h*area*(bur.thermo.T-T_w)*HeatOn
        return q
    
    def qfluxU(t): #Note - the only input to the passed through function is time
        rho=m_t/volume(t)
        T=(unb.mass*unb.thermo.T+bur.mass*bur.thermo.T)/(unb.mass+bur.mass)
        k=k_0*pow(T,n_k)        #W/m-K, for air
        mu=mu_0*pow(T,n_mu)      #kg/m-s, for air
        Re=rho*MPS*D/mu
        Nu=Nu_0*pow(Re,n_Nu)
        h=Nu*k/D
        area=surfA(t)*unb.volume/volume(t)
        q=h*area*(unb.thermo.T-T_w)/A_p*HeatOn
        return q
    
    def mflow(t): #Note - the only input to the passed through function is time
        theta=t*omega+IVC*np.pi/180 #TDC corresponds to theta=0
        if (theta >= theta_0r) and (theta <= theta_99r):
            # dxb=1/Deltathetar
            dxb=b*(m+1)/Deltathetar*pow(((theta-theta_0r)/Deltathetar),m)*math.exp(-b*pow(((theta-theta_0r)/Deltathetar),m+1))
        else:
            dxb=0
        mf=burnflag*m_t*omega*dxb    #mass flow unb to bur
    
        return mf
    
    ##  MAIN  - read inputs
    wb=load_workbook(fname)
    ws=wb.active
    r_c=ws['A1'].value      # compression ratio
    L=ws['A2'].value        # con rod length [m]
    D=ws['A3'].value        # bore [m]
    stroke=ws['A4'].value   # stroke [m]
    IVC=ws['A5'].value      # IVC crank angle where TDC = 0
    EVO=ws['A6'].value      # EVO crank angle where TDC = 0
    CrevOn=ws['A7'].value   # Flag to determine whether to use crevice model
    HeatOn=ws['A8'].value   # Flag to determine whether to use heat transfer model
    crevfrac=ws['A9'].value # crevice volume fraction of minimum volume
    RPM=ws['A10'].value     # engine speed
    b=ws['A11'].value       # Wiebe b parameter
    m=ws['A12'].value       # Wiebe m parameter
    theta_0=ws['A13'].value # Wiebe combustion start parameter
    Deltatheta=ws['A14'].value     # Wiebe combustion duration parameter
    eta_c=ws['A15'].value   # combustion efficiency
    T_IVC=ws['A16'].value   # IVC temperature [K]
    P_IVC=ws['A17'].value   # IVC pressure [bar]
    T_w=ws['A18'].value     # wall temperature [K]
    RON=ws['A19'].value     # fuel RON
    MON=ws['A20'].value     # fuel MON
    Phi=ws['A21'].value     # equivalence ratio
    burnflag=ws['A22'].value# fraction of mass transferred out of unburned zone 
    k_0=ws['A23'].value     # thermal conductivity [W/m-K]
    n_k=ws['A24'].value     # thermal conductivity temperature exponent
    mu_0=ws['A25'].value    # dynamic viscosity [kg/m-s]
    n_mu=ws['A26'].value    # viscosity temperature exponent
    Nu_0=ws['A27'].value    # Nusselt number correlation scaling constant
    n_Nu=ws['A28'].value    # Nusselt number correlation Reynolds number exponent
    Chemflag=ws['A29'].value# End gas chemistry flag; 1=on
    Y_EGR=ws['A30'].value   # EGR mass fraction
    Y_f_ref=ws['A31'].value # mass fraction of FUEL to the reformer
    Phi_ref=ws['A32'].value # reformer equivalence ratio
    Ref_comp=ws['A33'].value# reformer composition: 0=PCI; 1=equilibriu; 2=ideal
    mechanism=ws['A34'].value# kinetic mechanism
    
    ## Engine geometry calculations      
    a=stroke/2          # crank radius
    A_p=np.pi*D*D/4     # piston area [m^2]
    V_disp=A_p*stroke   # displacement volume [m^3]
    V_min=V_disp/(r_c-1)# TDC volume [m^3]
    V_max=V_min+V_disp  # BDC volume [m^3]
    MPS=2*stroke*RPM/60 #mean piston speed [m/s]
    
    ## Engine operating conditions
    theta_0r=theta_0*np.pi/180        #start of combustion in radians
    Deltathetar=Deltatheta*np.pi/180  #combustion duration in radians
    theta_99r=theta_0r+Deltathetar*math.pow(-math.log(0.01)/b,1/m)  #99% of mass burn for Wiebe
    omega=RPM*2*np.pi/60   #rotation rate in radian per second based on RPM
    
    T_0=T_IVC   # K
    P_0=1e5*P_IVC # Pa
    ct.suppress_thermo_warnings()
    env = ct.Solution('air.xml')
    env.TP = T_w, P_0
    gas = ct.Solution(mechanism)
    x_init=np.zeros(gas.X.shape)
    
    
    c4=ws['C41'].value #0.01      #n-butane
    c5=ws['C42'].value #0.0      #iso-pentane
    c5_2=ws['C43'].value #0.04   #n-pentane
    ic8=ws['C44'].value #0.93    #Iso-octane
    c6=ws['C45'].value #0.0      #added for 1-hexene
    nc7=ws['C46'].value #0.0    #n-heptane
    c7_2=ws['C47'].value #0.0    #tolune
    tmb=ws['C48'].value #0.02     #1,2,4 Trimethyl benzene 
    eth=ws['C49'].value #0       #Ethanol      


    
    x_init[gas.species_index('C4H10')]=c4
    x_init[gas.species_index('IC5H12')]=c5
    x_init[gas.species_index('NC5H12')]=c5_2
    x_init[gas.species_index('C6H12-1')]=c6
    x_init[gas.species_index('NC7H16')]=nc7
    x_init[gas.species_index('C6H5CH3')]=c7_2
    x_init[gas.species_index('IC8')]=ic8
    x_init[gas.species_index('T124MBZ')]=tmb
    x_init[gas.species_index('C2H5OH')]=eth
#    x_init[gas.species_index('NO')]=150e-6  #150 ppm NO

    
    ## Determine global composition
    nC=(4*c4)+(5*c5)+(5*c5_2)+(6*c6)+(7*nc7)+(8*ic8)+(9*tmb)+(2*eth)+(7*c7_2)
    nH=(10*c4)+(12*c5)+(12*c5_2)+(12*c6)+(16*nc7)+(18*ic8)+(12*tmb)+(6*eth)+(8*c7_2)
    nO=1*eth
    
    x_init[gas.species_index('O2')]=(nC+nH/4-nO/2)/Phi
    x_init[gas.species_index('N2')]=3.76*(nC+nH/4-nO/2)/Phi
    gas.TPX= 300, 1e5, x_init
    y_global=gas.Y
    

    
    ## Determine EGR composition
    keep=np.zeros(gas.X.shape)
    keep[gas.species_index('N2')]=1; keep[gas.species_index('O2')]=1; 
    keep[gas.species_index('CO2')]=1; keep[gas.species_index('CO')]=1; 
    keep[gas.species_index('H2O')]=1; keep[gas.species_index('H2')]=1; 
    gas.TPY=300, 1e5, y_global
    gas.equilibrate('HP')   #find equilibrium concentration
    y_EGR=gas.Y * keep;    # y_EGR[7:86]=0;  #set minor species to zero; based on MECHANISM
    gas.Y=y_EGR #use Cantera to renormalize mass fraction
    y_EGR=gas.Y   
    
    y_eng=1/(1+Y_EGR)*y_global++Y_EGR/(1+Y_EGR)*y_EGR
 
    gas.TPX=298, 1e5, x_init
    delta_uf=gas.standard_int_energies_RT*(ct.gas_constant*298)/gas.molecular_weights

    
    tmax=(EVO-IVC)/6/RPM # time for one revolution 
    step=8*(EVO-IVC)
    tim=np.linspace(tmax/step,tmax,step)
    theta=IVC+omega*tim*180/np.pi
    
    ## Cantera reactor setup
    gas.TPY = T_0, P_0, y_eng  #reactants
    # gas.set_multiplier(1)   #option to make reactants inert 
    unb=ct.IdealGasReactor(gas)
    if Chemflag ==0:
        unb.chemistry_enabled=False
    
    gas.TPY = T_0, P_0, y_eng  #products
    gas.equilibrate('HP')  #make the products hot so that mass will burn when it moves into products
    bur=ct.IdealGasReactor(gas)
    
    r3=ct.Reservoir(env)    #outside world
    
    gas.TPY = T_w, P_0, y_eng
    crev=ct.IdealGasReactor(gas)   #crevices
    crev.chemistry_enabled=False   #no reaction in crevices
    
    piston=ct.Wall(unb,r3,velocity=vel,A=A_p, Q=qfluxU)
    flame=ct.Wall(unb,bur, K=0.01, A=A_p)    #expansion rate K set to lowest value possible to have const P
    topland=ct.Wall(crev,r3, U=1e3) #heat xfer rate set to keep close to wall temperature
    head=ct.Wall(bur,r3, A=1, Q=qfluxB)
    
    mfc=ct.MassFlowController(unb, bur, mdot=mflow)
    V1=ct.Valve(unb,crev); V1.set_valve_coeff(1e-6*CrevOn)
    V2=ct.Valve(crev,bur); V2.set_valve_coeff(1e-6*CrevOn)
    
    sim = ct.ReactorNet([unb, bur, crev])
    sim.atol=1e-12
    sim.rtol=1e-6
    initvol=1e-3
    unb.volume=(1-initvol)*volume(0)
    bur.volume=initvol*volume(0)
    crev.volume=crevfrac*V_min
    
    m_t = unb.mass + bur.mass + crev.mass
    
    vol=np.zeros(step); T=np.zeros(step); vol2=np.zeros(step); P=np.zeros(step); 
    outdat=np.zeros((step,13))
    SpecMatrix=np.zeros((step,gas.X.shape[0]))
    #burnflag=0.82  #flag to stop combustion if unburned mass gets too low
#    y_old=unb.thermo.Y
    
    for i in range(step):#step):
        if math.fmod(i,32) == 0:
                   print(i/step)
        sim.advance(tim[i])
        # burnflag=unb.mass/m_t
        # if unb.mass/m_t < 1-eta_c:
        #     burnflag=0
        outdat[i,0]=IVC+tim[i]*RPM/60*360
        outdat[i,1]=bur.thermo.P/1e5
        outdat[i,2]=bur.thermo.T
        outdat[i,3]=bur.mass/m_t
        outdat[i,4]=qfluxB(tim[i])
        outdat[i,5]=unb.thermo.T
        outdat[i,6]=qfluxU(tim[i])*A_p
        outdat[i,7]=unb.mass/m_t
        outdat[i,8]=crev.mass/m_t
        outdat[i,9]=volume(sim.time)
        outdat[i,10]=-(unb.kinetics.net_production_rates*unb.thermo.molecular_weights).dot(delta_uf)*unb.volume*60/(RPM*360)
        outdat[i,11]=m_t
        outdat[i,12]=ct.gas_constant/gas.mean_molecular_weight
        outdat[i,13]=volume(tim[i])
        
        SpecMatrix[i]=gas.X
#        outdat[i,10]=-(unb.thermo.Y-y_old).dot(delta_uf)/(tim[2]-tim[1])*unb.mass
#        y_old=unb.thermo.Y
    
    
    return outdat, SpecMatrix

#    ## Read data
#    pf=np.zeros(1440); caf=np.zeros(1440)
#    with open('case210.csv') as csvfile:
#        readCSV = csv.reader(csvfile, delimiter=',')
#        i=0
#        for row in readCSV:
#            caf[i]=row[0]
#            pf[i]=row[1]
#            i=i+1
#            
#    ##Plot results        
#    plt.figure(1)
#    plt.plot(theta,outdat[:,1]/1e5,'k')
#    # plt.plot(theta,vol2/1e5,'g')
#    plt.plot(caf,pf/1e2,'r')
#    plt.xlim(-111,100)
#    # plt.figure(2)
#    # plt.plot(theta,T,'r')
#    # plt.plot(theta,vol,'b')
#    # plt.plot(theta,vol2,'r')
#    plt.show()
#    # plt.xlim(-15,15)
#    
#    np.savetxt('Crev_Heat.csv',outdat,delimiter=',');
